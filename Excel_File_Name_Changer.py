from openpyxl import load_workbook

def change_sheet_names(file_path, new_names):
              
    # Load workbook
    wb = load_workbook(file_path)

    # Getting list of existing sheet names
    existing_names = wb.sheetnames

    # Checking if the number of new names matches the number of existing sheets
    if len(existing_names) != len(new_names):
        print("Number of new names doesn't match the number of existing sheets.")
        return

    # Rename each sheet using the new names
    for sheet, new_name in zip(wb, new_names):
        sheet.title = new_name

    # Save the changes confirmed
    wb.save(file_path)
    print("Sheet names changed successfully.")

# Provide path to your Excel file
excel_file_path = ''

# Replace with new sheet names
new_sheet_names = []

#Calling function to change sheet names
change_sheet_names(excel_file_path, new_sheet_names)